import io
import json
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse
import openpyxl
from copy import copy
from openpyxl.styles import Font
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

app = FastAPI()

# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================
def copy_cell_with_formula_translation(src_cell, dst_cell):
    if src_cell.hyperlink:
        dst_cell.hyperlink = copy(src_cell.hyperlink)
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)

    if src_cell.value and isinstance(src_cell.value, str) and src_cell.value.startswith('='):
        try:
            translator = Translator(formula=src_cell.value, origin=src_cell.coordinate)
            dst_cell.value = translator.translate_formula(dst_cell.coordinate)
        except Exception:
            dst_cell.value = src_cell.value
    else:
        dst_cell.value = src_cell.value

def find_last_schedule_row(worksheet):
    for row_num in range(worksheet.max_row, 1, -1):
        cell_value = worksheet.cell(row=row_num, column=3).value
        if isinstance(cell_value, str) and "TOTAL" in cell_value.upper():
            return row_num
    return 30  # Fallback

def is_template_empty(worksheet, check_row, check_col):
    cell_value = worksheet.cell(row=check_row, column=check_col).value
    return cell_value is None or "N/A" in str(cell_value)

# ==============================================================================
# MAIN API ENDPOINT
# ==============================================================================
@app.post("/process-excel-panel/")
async def process_panel(
    panel_data_json: str = Form(...),
    file: UploadFile = File(...)
):
    if not file.filename.lower().endswith((".xlsm", ".xlsx")):
        raise HTTPException(status_code=400, detail="Invalid file format.")

    try:
        panel_data = json.loads(panel_data_json)
        contents = await file.read()
        wb_to_modify = openpyxl.load_workbook(io.BytesIO(contents), keep_vba=True)
        ws_to_modify = wb_to_modify.active
        pristine_template_wb = openpyxl.load_workbook('template.xlsm', keep_vba=True)
        pristine_ws = pristine_template_wb.active

        TEMPLATE_START_ROW = 7
        TEMPLATE_HEIGHT = 24
        TEMPLATE_COLUMN_COUNT = 44

        if is_template_empty(ws_to_modify, check_row=20, check_col=9):
            write_row = TEMPLATE_START_ROW
        else:
            last_total_row = find_last_schedule_row(ws_to_modify)
            insertion_row = last_total_row + 1
            ws_to_modify.insert_rows(insertion_row, amount=TEMPLATE_HEIGHT)

            for r_offset in range(TEMPLATE_HEIGHT):
                for c in range(1, TEMPLATE_COLUMN_COUNT + 1):
                    src_cell = pristine_ws.cell(row=TEMPLATE_START_ROW + r_offset, column=c)
                    dst_cell = ws_to_modify.cell(row=insertion_row + r_offset, column=c)
                    copy_cell_with_formula_translation(src_cell, dst_cell)

            for r_offset in range(TEMPLATE_HEIGHT):
                source_row_index = TEMPLATE_START_ROW + r_offset
                destination_row_index = insertion_row + r_offset
                if source_row_index in pristine_ws.row_dimensions:
                    ws_to_modify.row_dimensions[destination_row_index].height = pristine_ws.row_dimensions[source_row_index].height
            write_row = insertion_row

        for i in range(1, TEMPLATE_COLUMN_COUNT + 1):
            column_letter = get_column_letter(i)
            if column_letter in pristine_ws.column_dimensions:
                ws_to_modify.column_dimensions[column_letter].width = pristine_ws.column_dimensions[column_letter].width

        # --- Write Panel-Specific Data (with RCBO Highlighting) ---
        row = write_row
        ws_to_modify.cell(row=row, column=4).value = panel_data.get("panelName")
        ws_to_modify.cell(row=row + 6, column=7).value = panel_data.get("mountingType", "SURFACE")
        ws_to_modify.cell(row=row + 7, column=7).value = panel_data.get("ipDegree")

        link_cell = ws_to_modify.cell(row=row + 11, column=1)
        source_image_url = panel_data.get("sourceImageUrl")
        if source_image_url:
            link_cell.value = "panel image"
            link_cell.hyperlink = source_image_url
            link_cell.font = Font(color="0000FF", underline="single")

        recommendations = panel_data.get("recommendations", [])
        main_rec = next((r for r in recommendations if "MCCB" in r.get("breakerSpec", "")), None)
        if main_rec:
            cell = ws_to_modify.cell(row=row + 11, column=9)
            ref = main_rec.get("matchedPart", {}).get("Reference number", "")
            cell.value = ref
            if "RCBO" in main_rec.get("breakerSpec", ""):
                cell.font = Font(name="Montserrat", size=8, color="e50000")  # New red

        branch_recs = [r for r in recommendations if "MCCB" not in r.get("breakerSpec", "")]
        for i, rec in enumerate(branch_recs):
            current_row = row + 13 + i
            if current_row <= (row + TEMPLATE_HEIGHT - 1):
                ws_to_modify.cell(row=current_row, column=6).value = rec.get("quantity")
                cell = ws_to_modify.cell(row=current_row, column=9)
                ref = rec.get("matchedPart", {}).get("Reference number", "")
                cell.value = ref
                if "RCBO" in rec.get("breakerSpec", ""):
                    cell.font = Font(name="Montserrat", size=8, color="e50000")  # New red

        ws_to_modify.cell(row=row + 23, column=3).value = "TOTAL"

        # --- Clean Up Unused "OUTGOINGS" Rows ---
        outgoings_start_row = row + 13
        outgoings_end_row = row + 22
        for row_to_check in range(outgoings_end_row, outgoings_start_row - 1, -1):
            qty_cell = ws_to_modify.cell(row=row_to_check, column=6)
            part_num_cell = ws_to_modify.cell(row=row_to_check, column=9)
            part_num_is_default = (part_num_cell.value is None or "N/A" in str(part_num_cell.value))
            qty_is_default = (qty_cell.value == 1)
            if part_num_is_default and qty_is_default:
                ws_to_modify.delete_rows(row_to_check, 1)

        # --- Save and Return ---
        out = io.BytesIO()
        wb_to_modify.save(out)
        out.seek(0)

        original_filename = file.filename
        media_type = 'application/vnd.ms-excel.sheet.macroenabled.12' if original_filename.lower().endswith('.xlsm') else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return StreamingResponse(
            out,
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={original_filename}"}
        )

    except FileNotFoundError:
        raise HTTPException(status_code=500, detail="Server configuration error: The master template file is missing.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An error occurred during Excel processing: {str(e)}")
